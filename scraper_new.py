from selenium import webdriver
from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.options import Options
from selenium.webdriver.firefox.options import Options
from selenium.common.exceptions import NoSuchElementException
import openpyxl
import re


# enable headless mode in Selenium
options = Options()

# headless option for chrome
options.add_argument('--headless=new')

# headless option for firefox
options.add_argument('--headless')

# initialize an instance of the chrome/firefox driver (browser)
# it doesn't matter you write webdriver.Firefox or webdriver.Chrome, 
# you just need to drill options with what you import from selenium.webdriver.firefox.options.
driver_initial = webdriver.Chrome(
    options=options,
)

# visit your www.vehicle-spares.com site
driver_initial.get('https://www.vehicle-spares.com/collections/crash-parts?page=1')

# searching products...
product_count = driver_initial.find_element(By.CSS_SELECTOR, f"html body div.VehicleContainer div.ProdCount span")
product_num = int(re.search(r"\d+", ''.join(product_count.text.split(","))).group())
print(int(product_num),'items found.')

# # searching pages...
page_num = 346
print(346,'pages found.')

# to load the workbook with its path
Excelfile = openpyxl.load_workbook("Test.xlsx")
print("Excel sheet is loaded.")
# to identify active worksheet
ExcelSheet = Excelfile.active
print("active sheet is selected.")
# to identify maximum rows count
print ( ExcelSheet.max_row, 'rows')
# to identify maximum columns count
print ( ExcelSheet.max_column, 'columns')

item_from_sheet = ExcelSheet.cell(row=1, column=1).value
count = 0
while ExcelSheet.cell (row=count+1, column=2).value != None:
    count += 1
page = int((count -2) / 60) + 1
rest = count - ( page - 1 ) * 60 - 1
driver_initial.quit()
# # scraping search result


ExcelSheet.cell (row=1, column=1).value = "Product Page Link:"
ExcelSheet.cell (row=1, column=2).value = "Breadcrumb:"
ExcelSheet.cell (row=1, column=3).value = "Product Title:"
ExcelSheet.cell (row=1, column=4).value = "Product ID:"
ExcelSheet.cell (row=1, column=5).value = "Part No:"
ExcelSheet.cell (row=1, column=6).value = "inc vat:"
ExcelSheet.cell (row=1, column=7).value = "Condition:"
ExcelSheet.cell (row=1, column=8).value = "Quality:"
ExcelSheet.cell (row=1, column=9).value = "Make:"
ExcelSheet.cell (row=1, column=10).value = "Model:"
ExcelSheet.cell (row=1, column=11).value = "Body Style:"
ExcelSheet.cell (row=1, column=12).value = "Year Range:"
ExcelSheet.cell (row=1, column=13).value = "Approval or Side Fitment:"
ExcelSheet.cell (row=1, column=14).value = "Warranty:"
ExcelSheet.cell (row=1, column=15).value = "OE Equivalent:"

start_value = 0
stop_value = (page - 1) * 60
print(page, stop_value)
Excelfile.save ("Test.xlsx")
driver_main = webdriver.Chrome(
    options=options,
)
while page <= page_num:
    try:
        print('page', page,'starting')
        driver_main.get(f'https://www.vehicle-spares.com/collections/crash-parts?page={page}')
        index = 1
        product_link = driver_main.find_element(By.XPATH, f"//div[@class='grid-uniform PartsList']/div/div[{index}]/a").get_attribute('href')
        temp = []
        products_onscreen = len(driver_main.find_elements(By.XPATH, f"//div[@class='grid-uniform PartsList']/div/div"))
        print(products_onscreen)
        while index <= products_onscreen:
            product_link = driver_main.find_element(By.XPATH, f"//div[@class='grid-uniform PartsList']/div/div[{index}]/a").get_attribute('href')
            temp.append(product_link)
            ExcelSheet.cell (row = index + (page - 1) * 60 + rest + 1, column = 1).value = product_link
            print(index, 'products founded on the screen.')
            index += 1
        Excelfile.save ("Test.xlsx")
        start_value = stop_value + 1
        stop_value += len(temp)
        print('from',start_value+rest,'to', stop_value, 'in range of',start_value,'~', stop_value )
        index = rest
        rest = 0
        while index < products_onscreen:
            try:
                product_condition = "not specified"
                product_quality = "not specified"
                product_make = "not specified"
                product_model = "not specified"
                product_body_style = "not specified"
                product_year_range = "not specified"
                product_approval = "not specified"
                product_warranty = "not specified"
                product_oe_equivalent = "not specified"

                if index >= len(temp):
                    break
                else:
                    driver_main.get(temp[index])

                    bread_crumb = ' > '.join([a.text for a in driver_main.find_elements(By.XPATH, f"//main/div[1]/nav/a")[2:]])
                    ExcelSheet.cell (row = start_value + index + 1, column = 2).value = bread_crumb

                    product_title = driver_main.find_element(By.XPATH, f"//div[@class='ProductContainer']/div[@class='ProductRightColumn']/div[@class='ProductRightBlock']/div[@class='col-xs-12 ProductBlock']/h1[@class='product-title']").text
                    ExcelSheet.cell (row = start_value + index + 1, column = 3).value = product_title

                    product_id = driver_main.find_element(By.XPATH, f"//div[@class='col-xs-12 Pad0 ProductIdBar']/div[1]/span[@class='variant-sku']").text.split(": ")[1]
                    
                    ExcelSheet.cell (row = start_value + index + 1, column = 4).value = product_id

                    product_part_no = driver_main.find_element(By.XPATH, f"//div[@class='col-xs-12 Pad0 ProductIdBar']/div[2]").text
                    if product_part_no.find(':') == -1:
                        product_part_no = "not specified"
                    else:
                        product_part_no = product_part_no.split(": ")[1]
                    ExcelSheet.cell (row = start_value + index + 1, column = 5).value = product_part_no

                    product_inc_vat = driver_main.find_element(By.XPATH, f"//ul[@class='col-xs-12 Pad0 product-meta']/li/span[1]").text
                    ExcelSheet.cell (row = start_value + index + 1, column = 6).value = product_inc_vat

                    initialIndexList = ['Condition', 'Quality', 'Make', 'Model', 'Body Style', 'Year Range', 'Approval', 'Warranty', 'OE Equivalent']
                    initialList = [product_condition, product_quality, product_make, product_model, product_body_style, product_year_range, product_approval, product_warranty, product_oe_equivalent]
                    product_params = [th.text for tr in driver_main.find_elements(By.XPATH, f"//div[@class='ProductContainer']/div[6]/table/tbody/tr") for th in tr.find_elements(By.TAG_NAME, "th")]
                    for param in product_params:
                        if param in [x for x in initialIndexList]:
                            num = initialIndexList.index(param)
                            data_order = product_params.index(param)
                            initialList[num] = driver_main.find_element(By.XPATH, f"//div[@class='ProductContainer']/div[6]/table/tbody/tr[{data_order + 1}]/td").text
                        elif param == "Side Fitment" :
                            data_order = product_params.index(param)
                            initialList[6] = driver_main.find_element(By.XPATH, f"//div[@class='ProductContainer']/div[6]/table/tbody/tr[{data_order + 1}]/td").text

                    product_condition = initialList[0]
                    product_quality = initialList[1]
                    product_make = initialList[2]
                    product_model = initialList[3]
                    product_body_style = initialList[4]
                    product_year_range = initialList[5]
                    product_approval = initialList[6]
                    product_warranty = initialList[7]
                    product_oe_equivalent = initialList[8]

                    ExcelSheet.cell (row = start_value + index + 1, column = 7).value = product_condition
                    ExcelSheet.cell (row = start_value + index + 1, column = 8).value = product_quality
                    ExcelSheet.cell (row = start_value + index + 1, column = 9).value = product_make
                    ExcelSheet.cell (row = start_value + index + 1, column = 10).value = product_model
                    ExcelSheet.cell (row = start_value + index + 1, column = 11).value = product_body_style
                    ExcelSheet.cell (row = start_value + index + 1, column = 12).value = product_year_range
                    ExcelSheet.cell (row = start_value + index + 1, column = 13).value = product_approval
                    ExcelSheet.cell (row = start_value + index + 1, column = 14).value = product_warranty
                    ExcelSheet.cell (row = start_value + index + 1, column = 15).value = product_oe_equivalent
                    Excelfile.save ("Test.xlsx")

                    print(start_value + index, 'products scraped')
                    print(product_id)
                    index += 1
            except NoSuchElementException:
                break
        page += 1
    except NoSuchElementException:
        break

driver_main.quit()